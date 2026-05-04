"""File parsers — extract text representations for prompts.

Each parser returns plain text or markdown. Output is capped per-file
(MAX_PER_FILE_BYTES) by the caller; we don't truncate inside the parser.

Supported (per PROJECT_PLAN §14.2):
  - .txt, .md      → raw
  - .csv           → first 1000 rows as markdown table
  - .xlsx, .xls    → each sheet as markdown table
  - .pdf           → text via pypdf
  - code           → fenced markdown code block
"""

from __future__ import annotations

import csv
from pathlib import Path

from inclave_core.errors import CLIError
from inclave_core.workspace import kind_for

MAX_PER_FILE_BYTES = 100 * 1024  # 100 KB
MAX_TOTAL_BYTES = 200 * 1024  # 200 KB
MAX_FILES = 5

CSV_PREVIEW_ROWS = 1000


def parse(path: Path) -> str:
    kind = kind_for(path)
    if kind == "text":
        return _read_text(path)
    if kind == "csv":
        return _parse_csv(path)
    if kind == "xlsx":
        return _parse_xlsx(path)
    if kind == "pdf":
        return _parse_pdf(path)
    if kind == "code":
        return _parse_code(path)
    raise CLIError(
        f"unsupported file type: {path.suffix or '<no extension>'}. "
        "supported: .txt .md .csv .xlsx .xls .pdf and common code files"
    )


def _read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def _parse_code(path: Path) -> str:
    body = _read_text(path)
    lang = path.suffix.lstrip(".") or ""
    return f"```{lang}\n{body}\n```"


def _parse_csv(path: Path) -> str:
    with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.reader(f)
        try:
            header = next(reader)
        except StopIteration:
            return f"_(empty csv: {path.name})_"
        rows: list[list[str]] = []
        for i, row in enumerate(reader):
            if i >= CSV_PREVIEW_ROWS:
                break
            rows.append(row)
        truncated = False
        for _ in reader:
            truncated = True
            break

    out = [f"## CSV: {path.name} ({len(rows)} rows{' shown' if truncated else ''})"]
    out.append(_md_table(header, rows))
    if truncated:
        out.append(f"\n_(showing first {CSV_PREVIEW_ROWS} rows; more omitted)_")
    return "\n".join(out)


def _parse_xlsx(path: Path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_raw = next(rows_iter)
        except StopIteration:
            parts.append(f"## Sheet: {sheet_name} (empty)")
            continue
        header = ["" if c is None else str(c) for c in header_raw]
        rows: list[list[str]] = []
        n_rows = 0
        for row in rows_iter:
            n_rows += 1
            if n_rows > CSV_PREVIEW_ROWS:
                continue
            rows.append(["" if c is None else str(c) for c in row])
        parts.append(f"## Sheet: {sheet_name} ({n_rows} rows × {len(header)} cols)")
        parts.append(_md_table(header, rows))
        if n_rows > CSV_PREVIEW_ROWS:
            parts.append(f"\n_(showing first {CSV_PREVIEW_ROWS} rows of {n_rows})_")
    wb.close()
    return "\n\n".join(parts)


def _parse_pdf(path: Path) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    parts = [f"## PDF: {path.name} ({len(reader.pages)} pages)"]
    for i, page in enumerate(reader.pages, 1):
        text = page.extract_text() or ""
        if text.strip():
            parts.append(f"\n### page {i}\n{text.strip()}")
    return "\n".join(parts)


def _md_table(header: list[str], rows: list[list[str]]) -> str:
    if not header:
        return ""
    lines = ["| " + " | ".join(_cell(c) for c in header) + " |"]
    lines.append("| " + " | ".join("---" for _ in header) + " |")
    for row in rows:
        # pad to header width
        padded = row + [""] * (len(header) - len(row))
        lines.append("| " + " | ".join(_cell(c) for c in padded[: len(header)]) + " |")
    return "\n".join(lines)


def _cell(s: str) -> str:
    return s.replace("\n", " ").replace("|", "\\|")[:200]
